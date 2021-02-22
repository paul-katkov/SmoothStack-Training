# This program will work with any .xlsx file that exists within its directory and contains a month name seperated from the rest of the string by
# underscores ('_') and at least one 2- or a 4-digit sequence. If the file name contains a 2-digit sequence, the program will use it as 20xx. The program
# will always use the first digit sequence it finds, so if the file has multiple digit sequences, please make sure it is the first one, which references
# desired year.

# Certain information is assumed as known (e.g. the specific row/column index of the month list, the order of the promoter titles, etc.).

# It will work, however, even if the month list is out of chronological order (as long as the format is datetime and the list contains the date referenced
# in the file name). The length of the month list can also be configured via {data_offset} and {data_end_index} found below. Thank you!

from openpyxl import Workbook
import openpyxl
import logging
import sys

# Configuration:

logging.basicConfig(filename = "reports.log", format = "%(asctime)s [%(levelname)s]: %(message)s", level = logging.DEBUG)
logging.info(">SESSION START<")

months = ("january", "february", "march", "april", "may", "june", "july", "august", "september", "october", "november", "december")
# NOTE: Months indices referenced through this tuple will always be [0-11], but datetime indices are [1-12]

data_offset = 2 
# Actual data in the Excel file is preceded by two rows/columns [0, 1] of indices and titles
data_end_index = data_offset + 11
# Assumes a 12-month window (second sheet technically contains more than 12 entries, but only 12 different dates match between both sheets)

# Config End

def FindYear(name):
    year = ""

    for char in name:
        if char.isdigit():
            year += char
        elif len(year) == 2:
            return int(year) + 2000
        elif len(year) == 4:
            return int(year)
        else:
            year = ""
    
    if year == "":
        sys.exit()

def Load_Workbook(name):
    while True:
        try:
            excel = openpyxl.load_workbook(name)
        except:
            print("ERROR: Excel file was not found. Try again:")
            logging.error(f"Requested Excel file [ {name} ] was not found.")
            name = input()
            continue
        else:
            logging.info(f"File [ {name} ] loaded.")
            return excel, name

def GetMonthAndYear(name):
    try:
        month = months.index(list(set(name[:-5].split("_")).intersection(months))[0])
        year = FindYear(name)
    except IndexError:
        logging.error("Month was not detected in the file name.")
        sys.exit()
    except SystemExit:
        logging.error("Year was not detected in the file name.")
        sys.exit()
    except:
        logging.error("Requested file has invalid format or does not contain a month and/or year.")
        sys.exit()
    else:
        logging.info(f"Month and year detected in the file name ({months[month][0].upper()}{months[month][1:3]}-{str(year - 2000)}).")
        month = month + 1 # because datetime month indices start with 1 as opposed to the tuple at the top.
        return month, year

def SheetConfig(excel, month, year):
    sheets = excel.sheetnames

    try:
        summary_sheet = excel[sheets[0]]
        voc_sheet = excel[sheets[1]]
    except:
        logging.error(f"Requested Excel file does not have required format.")

    summary_row_index = 0
    voc_col_index = 0

    for month_row in summary_sheet.iter_cols(1, 1, data_offset, data_end_index):
        for date_cell in month_row:
            if date_cell.value.month == month and date_cell.value.year == year:
                summary_row_index = date_cell.row

    for month_col in voc_sheet.iter_rows(1, 1, data_offset, data_end_index):
        for date_cell in month_col:
            if date_cell.value.month == month and date_cell.value.year == year:
                voc_col_index = date_cell.column

    if summary_row_index == 0:
        logging.error(f"Requested month and/or year [{month}/{year}] were not found in the 'Summary Rolling MoM'.")
        sys.exit()
    elif voc_col_index == 0:
        logging.error(f"Requested month and/or year [{month}/{year}] was not found in the 'VOC Rolling MoM'.")
        sys.exit()
    else:
        return summary_sheet, voc_sheet, summary_row_index, voc_col_index

def ExtractDataIntoLog(month, year, summary_sheet, voc_sheet, summary_row_index, voc_col_index):
    logging.info(f"{summary_sheet.cell(1, 2).value.strip()}: {int(summary_sheet.cell(summary_row_index, 2).value)}")
    for i in range(3, 7):
        logging.info(f"{summary_sheet.cell(1, i).value.strip()}: {'{:.2f}'.format(summary_sheet.cell(summary_row_index, i).value * 100)}%")
    for u in range(4, 9, 2):
        promoter_group = voc_sheet.cell(u, 1).value.split()[0]
        amount = int(voc_sheet.cell(u, voc_col_index).value)
        goodbad = [f"is greater than 100. ({amount}) [BAD]", f"is less than 100. ({amount}) [GOOD]"]
        good = amount < 100
        if u == 4:
            goodbad = [f"is less than 200. ({amount}) [BAD]", f"is greater than 200. ({amount}) [GOOD]"]
            good = amount > 200
        elif u == 6:
            goodbad = [f"is less than 100. ({amount}) [BAD]", f"is greater than 100. ({amount}) [GOOD]"]
            good = amount > 100
        if good:
            logging.info(f"Number of {promoter_group} {goodbad[good]}")
        else:
            logging.warning(f"Number of {promoter_group} {goodbad[good]}")

def main():
    print("Please enter the name of .xlsx file you wish to access:")
    print("1. The name must contain a month name (e.g. 'september', 'march') seperated by underscore(s) ('_').")
    print("2. It must also contain a 2- or 4- digit sequence, referencing desired year.")
    print("3. In case of multiple sequences of digits, the program will utilize the first one as year.")
    print("4. If a 2-digit sequence is utilized, it will use 2000 as base (e.g. '04' -> '2004').")
    print("Lastly, please include .xlsx extension at the end of the file name.")
    print("EXAMPLES: expedia_report_monthly_january_2018.xlsx, june_17.xlsx, 2017_promoter_stats_august.xlsx\n")

    excelname = input()

    excel, excelname = Load_Workbook(excelname)

    try:
        month, year = GetMonthAndYear(excelname)
        summary_sheet, voc_sheet, summary_row_index, voc_col_index = SheetConfig(excel, month, year)
    except SystemExit:
        logging.warning("Program terminated prematurely.")
        logging.info(">SESSION OVER<\n")
    except:
        logging.error("Unknown error was detected.")
        logging.warning("Program terminated prematurely.")
        logging.info(">SESSION OVER<\n")
    else:
        ExtractDataIntoLog(month, year, summary_sheet, voc_sheet, summary_row_index, voc_col_index)
        logging.info(">SESSION OVER<\n")

if __name__ == '__main__':
    main()
